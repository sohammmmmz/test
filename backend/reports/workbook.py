"""Rendering a report as a real spreadsheet.

Not a CSV dump with a different extension. This is a document somebody forwards
to a stakeholder who will never open the app, so it has to answer on its own:
every sheet says what period it covers, dates are real dates rather than strings
so they sort and filter, percentages are real percentages so conditional
formatting works, and every table is frozen and filterable because a hundred
rows with no header pinned is a table nobody scrolls.

The palette is the app's, muted for print. A spreadsheet that arrives looking
like the screen it came from is easier to trust.
"""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# The app's blue hour, flattened for paper.
BRAND = "3D53E0"
INK = "14192B"
FAINT = "8590AA"
LINE = "D8DFEC"
DONE = "1F9D6B"
ATTENTION = "B8770F"
OVERDUE = "CF4740"
BAND = "F4F6FC"

HEAD_FILL = PatternFill("solid", fgColor=BRAND)
HEAD_FONT = Font(color="FFFFFF", bold=True, size=10.5)
TITLE_FONT = Font(color=INK, bold=True, size=15)
SUB_FONT = Font(color=FAINT, size=10)
BAND_FILL = PatternFill("solid", fgColor=BAND)
THIN = Side(style="thin", color=LINE)
CELL_BORDER = Border(bottom=THIN)


def _tone(value: str) -> Font | None:
    """Colour a verdict so a wall of numbers has somewhere for the eye to land."""
    return {
        "Over capacity": Font(color=OVERDUE, bold=True),
        "Full": Font(color=ATTENTION, bold=True),
        "Steady": Font(color=INK),
        "Spare capacity": Font(color=DONE),
        "Free": Font(color=FAINT),
        # "yes" only ever answers "is this slipping / overdue", so red is right.
        # Membership sync says synced/not synced for exactly that reason: reusing
        # yes and no there would have coloured a healthy row red.
        "yes": Font(color=OVERDUE, bold=True),
        "no": Font(color=FAINT),
        "synced": Font(color=FAINT),
        "not synced": Font(color=OVERDUE, bold=True),
    }.get(value)


def _sheet(book: Workbook, title: str, subtitle: str, columns: list[tuple[str, int]]):
    """A sheet with a title block, a header row, a freeze and a filter."""
    sheet = book.create_sheet(title)

    sheet["A1"] = title
    sheet["A1"].font = TITLE_FONT
    sheet["A2"] = subtitle
    sheet["A2"].font = SUB_FONT
    sheet.row_dimensions[1].height = 22
    sheet.row_dimensions[3].height = 6

    header = 4
    for index, (label, width) in enumerate(columns, start=1):
        cell = sheet.cell(row=header, column=index, value=label)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[header].height = 26

    # Frozen below the header, filters on it: a hundred rows with nothing pinned
    # is a table nobody scrolls.
    sheet.freeze_panes = sheet.cell(row=header + 1, column=1)
    sheet.auto_filter.ref = (
        f"A{header}:{get_column_letter(len(columns))}{header}"
    )
    return sheet, header


def _write(sheet, header: int, rows: list[list]):
    for offset, values in enumerate(rows):
        row_number = header + 1 + offset
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_number, column=index, value=value)
            cell.border = CELL_BORDER
            if offset % 2 == 1:
                cell.fill = BAND_FILL
            if isinstance(value, (date, datetime)):
                cell.number_format = "dd mmm yyyy"
                cell.alignment = Alignment(horizontal="left")
            elif isinstance(value, str):
                font = _tone(value)
                if font is not None:
                    cell.font = font
    # Widen the filter to cover the body, so it filters the data and not the
    # header alone.
    if rows:
        last = header + len(rows)
        sheet.auto_filter.ref = f"A{header}:{get_column_letter(len(rows[0]))}{last}"


def _percent(sheet, header: int, count: int, column: int):
    for row in range(header + 1, header + 1 + count):
        sheet.cell(row=row, column=column).number_format = '0"%"'


def build_workbook(report: dict) -> bytes:
    """The whole report as .xlsx bytes."""
    period = report["period"]
    summary = report["summary"]
    covers = f"{period.kind.title()} report · {period.label}"

    book = Workbook()
    book.remove(book.active)

    _summary_sheet(book, report, covers)
    _projects_sheet(book, report, covers)
    _people_sheet(book, report, covers)
    _assignments_sheet(book, report, covers)
    _milestones_sheet(book, report, covers)
    _activity_sheet(book, report, covers)

    book.properties.title = f"Morning Ledger — {covers}"
    book.properties.creator = summary["generated_by"]

    buffer = BytesIO()
    book.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------


def _summary_sheet(book: Workbook, report: dict, covers: str):
    summary = report["summary"]
    sheet = book.create_sheet("Summary")

    sheet["A1"] = "Morning Ledger"
    sheet["A1"].font = TITLE_FONT
    sheet["A2"] = covers
    sheet["A2"].font = SUB_FONT
    sheet["A3"] = (
        f"Prepared for {summary['generated_by']} · "
        f"{summary['generated_at'].strftime('%d %b %Y, %H:%M')}"
    )
    sheet["A3"].font = SUB_FONT

    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 62

    blocks = [
        ("Projects", None, None),
        ("Projects tracked", summary["projects"], "Every project you own."),
        ("Active", summary["active_projects"], "Status set to active."),
        ("Slipping", summary["slipping"], "Past a milestone due date."),
        ("Not fully set up", summary["not_ready"],
         "Missing a dated milestone, a document, or people."),
        (None, None, None),
        ("People", None, None),
        ("People covered", summary["people"], "Everyone on your teams, and you."),
        ("Over capacity", summary["over_capacity"],
         f"Holding more than {summary['capacity_basis']} open items, overdue counted twice."),
        ("Nothing assigned", summary["idle"], "No open tasks and no open todos."),
        (None, None, None),
        ("Work", None, None),
        ("Open tasks", summary["open_tasks"], "GitLab issues assigned and not closed."),
        ("Overdue tasks", summary["overdue_tasks"], "Open and past their due date."),
        ("Tasks closed this period", summary["tasks_closed"], "Closed inside the window."),
        ("Todos closed this period", summary["todos_closed"],
         "Closed by an owner in the morning meeting."),
        ("Todos awaiting closing", summary["todos_awaiting"],
         "Marked done by the person and not yet confirmed."),
        ("Morning meetings held", summary["meetings_held"], "Completed rounds."),
    ]

    row = 5
    for label, value, note in blocks:
        if label is None:
            row += 1
            continue
        cell = sheet.cell(row=row, column=1, value=label)
        if value is None:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = HEAD_FILL
            sheet.cell(row=row, column=2).fill = HEAD_FILL
            sheet.cell(row=row, column=3).fill = HEAD_FILL
        else:
            cell.font = Font(color=INK)
            cell.border = CELL_BORDER
            number = sheet.cell(row=row, column=2, value=value)
            number.font = Font(bold=True, size=12, color=INK)
            number.alignment = Alignment(horizontal="right")
            number.border = CELL_BORDER
            hint = sheet.cell(row=row, column=3, value=note)
            hint.font = SUB_FONT
            hint.border = CELL_BORDER
        row += 1

    row += 1
    sheet.cell(row=row, column=1, value="How bandwidth is worked out").font = Font(
        color=INK, bold=True
    )
    sheet.cell(row=row + 1, column=1, value=(
        "Open GitLab tasks plus open todos, with overdue work counted twice, "
        f"against a nominal {summary['capacity_basis']} items per person. "
        "It is a heuristic for spotting who to talk to, not a measure of anyone."
    )).font = SUB_FONT
    sheet.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=3)
    sheet.cell(row=row + 1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    sheet.row_dimensions[row + 1].height = 34


def _projects_sheet(book: Workbook, report: dict, covers: str):
    columns = [
        ("Project", 26), ("Status", 12), ("Repository", 30), ("Team", 16),
        ("People", 8), ("Progress", 10), ("Tasks done", 11), ("Tasks", 8),
        ("Closed this period", 12), ("Milestones", 11), ("Milestones closed", 12),
        ("Overdue milestones", 12), ("Slipping", 10), ("Next milestone", 24),
        ("Next due", 13), ("Started", 13), ("Target end", 13),
        ("Set up", 9), ("Still missing", 40),
    ]
    sheet, header = _sheet(book, "Projects", covers, columns)
    rows = [[
        p["name"], p["status"], p["repository"], p["team"], p["people"],
        p["percent"], p["tasks_done"], p["tasks"], p["closed_in_period"],
        p["milestones"], p["milestones_closed"], p["overdue_milestones"],
        "yes" if p["is_slipping"] else "no", p["next_milestone"], p["next_due"],
        p["started_on"], p["target_end_on"], p["setup"], p["missing"],
    ] for p in report["projects"]]
    _write(sheet, header, rows)
    _percent(sheet, header, len(rows), 6)


def _people_sheet(book: Workbook, report: dict, covers: str):
    columns = [
        ("Person", 22), ("GitLab", 18), ("Department", 15), ("Role", 9),
        ("Bandwidth", 15), ("Load", 9), ("Open tasks", 10), ("Overdue", 9),
        ("Open todos", 10), ("Carrying over", 12), ("Sitting for days", 13),
        ("Tasks closed", 11), ("Todos in period", 12), ("Todos closed", 11),
        ("Awaiting closing", 13), ("Projects", 9), ("Working on", 44),
    ]
    sheet, header = _sheet(book, "People", covers, columns)
    rows = [[
        p["name"], p["gitlab_username"], p["department"], p["role"],
        p["bandwidth"], p["bandwidth_percent"], p["open_tasks"], p["overdue_tasks"],
        p["todos_open_today"], p["todos_carrying"], p["todos_stale"],
        p["tasks_closed_in_period"], p["todos_in_period"], p["todos_closed"],
        p["todos_awaiting"], p["projects"], p["project_names"],
    ] for p in report["people"]]
    _write(sheet, header, rows)
    _percent(sheet, header, len(rows), 6)


def _assignments_sheet(book: Workbook, report: dict, covers: str):
    columns = [
        ("Project", 26), ("Status", 12), ("Person", 22), ("Department", 15),
        ("Branch", 24), ("On GitLab", 10), ("Open tasks", 10), ("Overdue", 9),
        ("Closed this period", 14),
    ]
    sheet, header = _sheet(book, "Who is where", covers, columns)
    _write(sheet, header, [[
        a["project"], a["project_status"], a["person"], a["department"],
        a["branch"], a["on_gitlab"], a["open_tasks"], a["overdue_tasks"],
        a["closed_in_period"],
    ] for a in report["assignments"]])


def _milestones_sheet(book: Workbook, report: dict, covers: str):
    columns = [
        ("Project", 26), ("Milestone", 30), ("State", 10), ("Start", 13),
        ("Due", 13), ("Days left", 10), ("Overdue", 9), ("Progress", 10),
        ("Tasks done", 11), ("Tasks", 8), ("Closed this period", 14),
    ]
    sheet, header = _sheet(book, "Milestones", covers, columns)
    rows = [[
        m["project"], m["milestone"], m["state"], m["start_date"], m["due_date"],
        m["days_remaining"], "yes" if m["is_overdue"] else "no", m["percent"],
        m["tasks_done"], m["tasks"], m["closed_in_period"],
    ] for m in report["milestones"]]
    _write(sheet, header, rows)
    _percent(sheet, header, len(rows), 8)


def _activity_sheet(book: Workbook, report: dict, covers: str):
    columns = [
        ("Date", 14), ("Day", 12), ("Working day", 12), ("Todos set", 11),
        ("Closed", 9), ("Awaiting closing", 14), ("Still open", 11),
        ("Tasks closed", 12), ("Meetings held", 12),
    ]
    sheet, header = _sheet(book, "Day by day", covers, columns)
    _write(sheet, header, [[
        a["date"], a["weekday"], "yes" if a["is_working_day"] else "no",
        a["todos"], a["closed"], a["awaiting"], a["still_open"],
        a["tasks_closed"], a["meetings"],
    ] for a in report["activity"]])
